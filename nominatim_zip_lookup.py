import openpyxl
import requests
import time
from openpyxl import load_workbook

EXCEL_PATH = 'list 1.xlsx'
OUTPUT_PATH = 'list 1 with zip.xlsx'
USER_AGENT = 'AberrantZipLookup/1.0 (your_email@example.com)'


def get_zipcode_nominatim(street, city, state):
    address = f"{street}, {city}, {state}, USA"
    url = 'https://nominatim.openstreetmap.org/search'
    params = {
        'q': address,
        'format': 'json',
        'addressdetails': 1,
        'limit': 1
    }
    headers = {'User-Agent': USER_AGENT}
    try:
        response = requests.get(url, params=params, headers=headers, timeout=10)
        response.raise_for_status()
        data = response.json()
        if data and 'address' in data[0] and 'postcode' in data[0]['address']:
            return data[0]['address']['postcode']
        else:
            return ''
    except Exception as e:
        print(f"Error for {address}: {e}")
        return ''

def main():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    street_col = headers.get('Street 1')
    city_col = headers.get('City')
    state_col = headers.get('State')
    if not (street_col and city_col and state_col):
        print('Missing required columns: Street 1, City, State')
        return
    zip_col = headers.get('Zip Code')
    if not zip_col:
        ws.cell(row=1, column=ws.max_column + 1, value='Zip Code')
        zip_col = ws.max_column
    for row in range(2, ws.max_row + 1):
        street = ws.cell(row=row, column=street_col).value or ''
        city = ws.cell(row=row, column=city_col).value or ''
        state = ws.cell(row=row, column=state_col).value or ''
        if street and city and state:
            zip_code = get_zipcode_nominatim(street, city, state)
            ws.cell(row=row, column=zip_col, value=zip_code)
            print(f"{street}, {city}, {state} -> {zip_code}")
            time.sleep(1)  # Be polite to the API
    wb.save(OUTPUT_PATH)
    print(f'Updated file saved as {OUTPUT_PATH}')

if __name__ == '__main__':
    main()
