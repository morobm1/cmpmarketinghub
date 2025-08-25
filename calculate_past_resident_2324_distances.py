import openpyxl
import requests
import time
from math import radians, sin, cos, sqrt, atan2

def find_header_row(ws, must_have):
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and all(h in row for h in must_have):
            return row_idx, list(row)
    raise Exception("Header row with required columns not found.")

UNIVERSITY_ADDR = '18612 Beardslee Blvd, Bothell, WA 98011'
USER_AGENT = 'AberrantZipLookup/1.0 (your_email@example.com)'
EXCEL_PATH = 'Past resident 2324.xlsx'
OUTPUT_PATH = 'Past_resident_2324_with_distances.xlsx'

must_have = ['Street Line1', 'City', 'State/Province', 'Postal Code']

# Haversine function
def haversine(lat1, lon1, lat2, lon2):
    R = 3958.8  # miles
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1)*cos(lat2)*sin(dlon/2)**2
    c = 2*atan2(sqrt(a), sqrt(1 - a))
    return R * c

# Geocode address
def geocode(address):
    url = 'https://nominatim.openstreetmap.org/search'
    params = {'q': address, 'format': 'json', 'limit': 1}
    headers = {'User-Agent': USER_AGENT}
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        if data:
            return float(data[0]['lat']), float(data[0]['lon'])
    except Exception as e:
        print(f"Geocode error for '{address}': {e}")
    return None, None

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    header_row_idx, headers = find_header_row(ws, must_have)
    # Build colmap
    colmap = {name: idx + 1 for idx, name in enumerate(headers)}
    # Prepare output columns
    lat_col = len(headers) + 1
    lon_col = len(headers) + 2
    dist_col = len(headers) + 3
    ws.cell(row=header_row_idx, column=lat_col, value='Latitude')
    ws.cell(row=header_row_idx, column=lon_col, value='Longitude')
    ws.cell(row=header_row_idx, column=dist_col, value='Distance to Univ (mi)')

    uni_lat, uni_lon = geocode(UNIVERSITY_ADDR)
    assert uni_lat is not None and uni_lon is not None, 'University address geocoding failed.'

    for row in range(header_row_idx+1, ws.max_row+1):
        s1 = ws.cell(row=row, column=colmap.get('Street Line1',0)).value or ''
        s2 = ws.cell(row=row, column=colmap.get('Street Line2',0)).value or ''
        s3 = ws.cell(row=row, column=colmap.get('Street Line3',0)).value or ''
        city = ws.cell(row=row, column=colmap.get('City',0)).value or ''
        state = ws.cell(row=row, column=colmap.get('State/Province',0)).value or ''
        zipc = ws.cell(row=row, column=colmap.get('Postal Code',0)).value or ''
        country = ws.cell(row=row, column=colmap.get('Country',0)).value or 'USA'
        address = ', '.join(filter(None, [str(s1), str(s2), str(s3), str(city), str(state), str(zipc), str(country)]))
        lat, lon = geocode(address)
        if lat is not None and lon is not None:
            dist = haversine(lat, lon, uni_lat, uni_lon)
            ws.cell(row=row, column=lat_col, value=lat)
            ws.cell(row=row, column=lon_col, value=lon)
            ws.cell(row=row, column=dist_col, value=round(dist,2))
            print(f"{address} -> {round(dist,2)} mi")
        else:
            ws.cell(row=row, column=lat_col, value='')
            ws.cell(row=row, column=lon_col, value='')
            ws.cell(row=row, column=dist_col, value='')
        time.sleep(1)

    wb.save(OUTPUT_PATH)
    print(f'Saved: {OUTPUT_PATH}')

if __name__ == '__main__':
    main()
