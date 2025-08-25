import openpyxl
import csv

EXCEL_PATH = 'Past_resident_with_distances.xlsx'
CSV_PATH = 'Past_resident_with_distances.csv'

must_have = ['Street Line1', 'City', 'State/Province', 'Postal Code']

def find_header_row(ws, must_have):
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and all(h in row for h in must_have):
            return row_idx, list(row)
    raise Exception("Header row with required columns not found.")

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active
header_row_idx, headers = find_header_row(ws, must_have)

with open(CSV_PATH, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(headers + ['Latitude', 'Longitude', 'Distance to Univ (mi)'])
    for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
        if any(row):
            writer.writerow(row)

print(f'Converted {EXCEL_PATH} to {CSV_PATH}')
