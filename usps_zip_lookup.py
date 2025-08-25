import openpyxl
import requests
from openpyxl import load_workbook
from xml.etree import ElementTree as ET

# INSTRUCTIONS:
# 1. Install dependencies: pip install openpyxl requests
# 2. Insert your USPS USERID below
# 3. Run this script: python usps_zip_lookup.py

USPS_USERID = '8CAPST87O7389'  # <-- Replace with your USPS Web Tools USERID
EXCEL_PATH = 'list 1.xlsx'
OUTPUT_PATH = 'list 1 with zip.xlsx'


def get_zipcode(street, city, state):
    url = 'https://secure.shippingapis.com/ShippingAPI.dll'
    xml = f'''<AddressValidateRequest USERID="{USPS_USERID}"><Revision>1</Revision><Address ID="0"><Address1></Address1><Address2>{street}</Address2><City>{city}</City><State>{state}</State></Address></AddressValidateRequest>'''
    params = {
        'API': 'Verify',
        'XML': xml
    }
    try:
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        tree = ET.fromstring(response.content)
        zip5 = tree.find('.//Zip5')
        if zip5 is not None and zip5.text:
            return zip5.text
        else:
            return ''
    except Exception as e:
        print(f"Error for {street}, {city}, {state}: {e}")
        return ''

def main():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    # Find column indices
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    street_col = headers.get('Street 1')
    city_col = headers.get('City')
    state_col = headers.get('State')
    if not (street_col and city_col and state_col):
        print('Missing required columns: Street 1, City, State')
        return
    # Add Zip Code column if not present
    zip_col = headers.get('Zip Code')
    if not zip_col:
        ws.cell(row=1, column=ws.max_column + 1, value='Zip Code')
        zip_col = ws.max_column
    # Process each row
    for row in range(2, ws.max_row + 1):
        street = ws.cell(row=row, column=street_col).value or ''
        city = ws.cell(row=row, column=city_col).value or ''
        state = ws.cell(row=row, column=state_col).value or ''
        if street and city and state:
            zip_code = get_zipcode(street, city, state)
            ws.cell(row=row, column=zip_col, value=zip_code)
    wb.save(OUTPUT_PATH)
    print(f'Updated file saved as {OUTPUT_PATH}')

if __name__ == '__main__':
    main()
