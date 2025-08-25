import openpyxl
import requests
import time
from math import radians, sin, cos, sqrt, atan2
from openpyxl import load_workbook

EXCEL_PATH = 'Bothell - FirstYearAdmitAndDeposit.xlsx'
OUTPUT_PATH = 'Bothell_with_distances.xlsx'
USER_AGENT = 'AberrantZipLookup/1.0 (your_email@example.com)'

UNIVERSITY_ADDR = '18612 Beardslee Blvd, Bothell, WA 98011'

# Haversine function for distance calculation
def haversine(lat1, lon1, lat2, lon2):
    R = 3958.8  # Earth radius in miles
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1)*cos(lat2)*sin(dlon/2)**2
    c = 2*atan2(sqrt(a), sqrt(1 - a))
    return R * c

# Geocode address using Nominatim
def geocode(address):
    url = 'https://nominatim.openstreetmap.org/search'
    params = {'q': address, 'format': 'json', 'limit': 1}
    headers = {'User-Agent': USER_AGENT}
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        if data:
            return float(data[0]['lat']), float(data[0]['lon'])
    except Exception as e:
        print(f"Geocode error for '{address}': {e}")
    return None, None

def main():
    # Geocode university location
    uni_lat, uni_lon = geocode(UNIVERSITY_ADDR)
    assert uni_lat is not None and uni_lon is not None, 'University address geocoding failed.'

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    
    street1_col = headers.get('Street 1')
    street2_col = headers.get('Street 2')
    city_col = headers.get('City')
    state_col = headers.get('State')
    zip_col = headers.get('ZipCode')
    # Prepare output columns
    lat_col = ws.max_column + 1
    lon_col = ws.max_column + 2
    dist_col = ws.max_column + 3
    ws.cell(row=1, column=lat_col, value='Latitude')
    ws.cell(row=1, column=lon_col, value='Longitude')
    ws.cell(row=1, column=dist_col, value='Distance to Univ (mi)')

    for row in range(2, ws.max_row + 1):
        street1 = ws.cell(row=row, column=street1_col).value or ''
        street2 = ws.cell(row=row, column=street2_col).value or ''
        city = ws.cell(row=row, column=city_col).value or ''
        state = ws.cell(row=row, column=state_col).value or ''
        zipc = ws.cell(row=row, column=zip_col).value or ''
        address = ', '.join(filter(None, [str(street1), str(street2), str(city), str(state), str(zipc), 'USA']))
        lat, lon = geocode(address)
        if lat is not None and lon is not None:
            dist = haversine(lat, lon, uni_lat, uni_lon)
            ws.cell(row=row, column=lat_col, value=lat)
            ws.cell(row=row, column=lon_col, value=lon)
            ws.cell(row=row, column=dist_col, value=round(dist, 2))
            print(f"{address} -> {round(dist,2)} mi")
        else:
            ws.cell(row=row, column=lat_col, value='')
            ws.cell(row=row, column=lon_col, value='')
            ws.cell(row=row, column=dist_col, value='')
        time.sleep(1)  # polite use for Nominatim
    wb.save(OUTPUT_PATH)
    print(f'Saved with distances: {OUTPUT_PATH}')

if __name__ == '__main__':
    main()
