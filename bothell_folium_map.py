import folium
import openpyxl

EXCEL_PATH = 'Bothell_with_distances.xlsx'
MAP_PATH = 'bothell_folium_map.html'

UNIVERSITY_COORDS = (47.758284, -122.191377)
UNIVERSITY_ADDR = 'University of Bothell'  # Will show in popup
RADIUS_MI = 10

# Read data from Excel
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active
headers = [cell.value for cell in ws[1]]
hdr_idx = {header: i for i, header in enumerate(headers)}

# Prepare the map
folium_map = folium.Map(location=UNIVERSITY_COORDS, zoom_start=10, tiles='OpenStreetMap')

# University marker and radius
folium.Marker(
    UNIVERSITY_COORDS,
    popup=f"<b>{UNIVERSITY_ADDR}</b><br>18612 Beardslee Blvd, Bothell, WA",
    icon=folium.Icon(color='blue', icon='university', prefix='fa')
).add_to(folium_map)

folium.Circle(
    location=UNIVERSITY_COORDS,
    radius=RADIUS_MI * 1609.34,
    color="#1c75bc",
    fill=True,
    fill_opacity=0.18
).add_to(folium_map)

# Plot each address
for row in ws.iter_rows(min_row=2, values_only=True):
    lat = row[hdr_idx.get('Latitude')]
    lon = row[hdr_idx.get('Longitude')]
    dist = row[hdr_idx.get('Distance to Univ (mi)')]
    if lat is None or lon is None:
        continue
    addr = ', '.join(str(row[hdr_idx.get(h)]) for h in ['Street 1','Street 2','City','State','ZipCode'] if row[hdr_idx.get(h)])
    color = 'green' if dist is not None and float(dist) <= RADIUS_MI else 'red'
    folium.CircleMarker(
        location=(lat, lon),
        radius=5,
        color=color,
        fill=True,
        fill_color=color,
        fill_opacity=0.85,
        popup=f"<b>{addr}</b><br>Dist: {dist} mi"
    ).add_to(folium_map)

folium_map.save(MAP_PATH)
print(f"Wrote: {MAP_PATH}")
